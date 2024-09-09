import csv

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
import pandas as pd
from .models import Matiere, Note, Semestre, Unite, UniteClasse
from inscription.models import AnneeAcademique, Classe, Etudiant, Inscription
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

def getUnites(request):
    classe_id = request.GET.get('classe')
    annee_id = request.GET.get('annee_academique')
    semestre_id = request.GET.get('semestre')

    try:
        # Filtrer par les relations à partir de UniteClasse
        unite_classes = UniteClasse.objects.filter(
            annee_id=annee_id,  # Filtre par année académique
            classe_id=classe_id,  # Filtre par classe
            semestre_id=semestre_id  # Filtre par semestre
        )


        # Obtenir les unités associées
        unites = unite_classes.select_related('unite').values_list('unite__nom', 'unite__id')
        
   
        return JsonResponse(list(unites), safe=False)

    except Exception as e:
        print(f"Erreur lors de la récupération des unités : {e}")
        return JsonResponse({'error': 'Erreur lors de la récupération des unités'}, status=500)


def getMatieres(request):
    classe_id = request.GET.get('classe_id')
    annee_id = request.GET.get('annee_id')
    semestre_id = request.GET.get('semestre_id')
    unite_id = request.GET.get('unite_id')

    try:
        # Filtrer les unités qui correspondent à la classe, l'année et le semestre donnés
        unite_classe = UniteClasse.objects.filter(
            classe_id=classe_id,
            annee_id=annee_id,
            semestre_id=semestre_id,
            unite_id=unite_id
        ).first()

        # Vérifiez si l'unité correspondante existe
        if not unite_classe:
            return JsonResponse({'error': 'Aucune unité trouvée pour les critères spécifiés.'}, status=404)

        # Obtenir les matières associées à l'unité
        matieres = Matiere.objects.filter(unite=unite_classe.unite).values_list('nom', 'id')
        
        print(f"Liste des matières: {list(matieres)}")

        return JsonResponse(list(matieres), safe=False)

    except Exception as e:
        print(f"Erreur lors de la récupération des matières : {e}")
        return JsonResponse({'error': 'Erreur lors de la récupération des matières'}, status=500)

def exportReleveNote(request):
    annees = AnneeAcademique.objects.all()
    classes = Classe.objects.all()
    semestres = Semestre.objects.all()

    context = {
        'annees': annees,
        'classes': classes,
        'semestres': semestres,
    }

    if request.method == 'POST':
        annee_academique_id = request.POST.get('annee_academique')
        classe_id = request.POST.get('classe')
        semestre_id = request.POST.get('semestre')
        unite_id = request.POST.get('unite')
        matiere_id = request.POST.get('matiere')

        annee = AnneeAcademique.objects.filter(pk=annee_academique_id).first()
        classe = Classe.objects.filter(pk=classe_id).first()
        semestre = Semestre.objects.filter(pk=semestre_id).first()
        unite = Unite.objects.filter(pk=unite_id).first()
        matiere = Matiere.objects.filter(pk=matiere_id).first()

        etudiants = Inscription.get_etudiants_inscrits(annee_academique_id, classe_id).values_list('matricule', 'nom', 'prenom')

        # Créer un nouveau fichier Excel
        wb = Workbook()
        ws = wb.active

        # Ajouter un titre
        title = f"ETUDIANTS INSCRITS. ANNEE:{annee.libelle}, CLASSE: {classe.libelle}, SEMESTRE: {semestre.nom},  UNITE: {unite.nom}, MATIERE: {matiere.nom}"
        max_column = 10  # Nombre total de colonnes (changez ceci si vous avez plus de colonnes)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
        ws['A1'] = title
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Ajuster la hauteur de la ligne pour le titre
        ws.row_dimensions[1].height = 40  # Vous pouvez ajuster la hauteur selon vos besoins

        # Ajouter des entêtes
        headers = ['matricule', 'nom', 'prenom', 'Note']
        ws.append(headers)

        # Style des entêtes
        header_font = Font(bold=True, color='FFFFFF')
        fill = PatternFill(start_color="000000FF", end_color="000000FF", fill_type="solid")
        for cell in ws[2]:  # 2ème ligne (après le titre)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')

        # Style des bordures
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Ajouter les données des étudiants
        for etudiant in etudiants:
            ws.append(list(etudiant) + [''])  # Ajouter une colonne vide pour "Note"

        # Appliquer des bordures aux cellules contenant des données
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

        # Définir la réponse HTTP avec le type de contenu approprié
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="etudiants-inscrits.xlsx"'

        # Enregistrer le fichier Excel dans la réponse
        wb.save(response)

        return response
    else:
        return render(request, 'note/releves_note.html', context)
    


def importerNote(request):
    annees = AnneeAcademique.objects.all()
    classes = Classe.objects.all()
    semestres = Semestre.objects.all()

    liste_notes = []

    context = {
        'annees': annees,
        'classes': classes,
        'semestres': semestres,
    }

    if request.method == 'POST' and request.FILES['fichier']:
        
        annee = request.POST.get('annee_academique')
        classe = request.POST.get('classe')
        semestre = request.POST.get('semestre')
        unite = request.POST.get('unite')
        matiere = request.POST.get('matiere')

        matiere_id = Matiere.objects.get(id=matiere)
        annee_id = AnneeAcademique.objects.get(id=annee)

        fichier = request.FILES['fichier']
        df = pd.read_excel(fichier)



        for index, row in df.iterrows():
            try:
                etudiant_one = {}
                etudiant = Etudiant.objects.get(matricule=row['MATRICULE'])  # ou autre identifiant unique
                note = row['NOTE']
                
                Note.objects.get_or_create(etudiant=etudiant, valeur=note, matiere=matiere_id, annee=annee_id)
                etudiant_one['MATRICULE'] = row['MATRICULE']
                etudiant_one['NOM'] = row['NOM']
                etudiant_one['PRENOM'] = row['PRENOM']
                etudiant_one['NOTE'] = note
                liste_notes.append(etudiant_one)
                
            except Etudiant.DoesNotExist:
                print("Etudiant non trouvé!")
        
        context = {
            'alert_message': 'Vos notes ont été importées avec succès !',
            'liste_notes': liste_notes
        }
        return render(request, 'note/importer_notes.html', context)
    else:
        return render(request, 'note/importer_notes.html', context)